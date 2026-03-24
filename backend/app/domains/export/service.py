from io import BytesIO

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.asset import Asset
from app.models.asset_type import AssetType
from app.models.organization import Organization
from app.models.project import Project
from app.models.currency import Currency
from app.models.renewal_period import RenewalPeriod
from app.domains.export.schemas import ExportAssetsRequest

DEFAULT_COLUMNS = [
    "id",
    "name",
    "type",
    "organization",
    "project",
    "cost",
    "currency",
    "purchase_date",
    "next_payment_date",
    "renewal_type",
    "admin_account",
    "comment",
]

COLUMN_HEADERS = {
    "id": "ID",
    "name": "Наименование",
    "type": "Тип",
    "organization": "Организация",
    "project": "Проект",
    "cost": "Стоимость",
    "currency": "Валюта",
    "purchase_date": "Дата покупки",
    "next_payment_date": "Следующий платёж",
    "renewal_type": "Тип продления",
    "admin_account": "Админ. учётная запись",
    "comment": "Комментарий",
}

RENEWAL_TYPE_LABELS = {
    "fixed": "Фиксированный",
    "manual": "Плавающий",
}


async def _load_lookups(db: AsyncSession) -> dict:
    types = {r.id: r.name for r in (await db.execute(select(AssetType))).scalars().all()}
    orgs = {r.id: r.name for r in (await db.execute(select(Organization))).scalars().all()}
    projects = {r.id: r.name for r in (await db.execute(select(Project))).scalars().all()}
    currencies = {
        r.id: f"{r.code} ({r.symbol})"
        for r in (await db.execute(select(Currency))).scalars().all()
    }
    return {
        "type": types,
        "organization": orgs,
        "project": projects,
        "currency": currencies,
    }


def _get_cell_value(asset: Asset, col: str, lookups: dict) -> object:
    if col == "type":
        return lookups["type"].get(asset.type_id, asset.type_id)
    if col == "organization":
        return lookups["organization"].get(asset.organization_id, asset.organization_id)
    if col == "project":
        return lookups["project"].get(asset.project_id, "") if asset.project_id else ""
    if col == "currency":
        return lookups["currency"].get(asset.currency_id, asset.currency_id)
    if col == "renewal_type":
        return RENEWAL_TYPE_LABELS.get(asset.renewal_type, asset.renewal_type)

    value = getattr(asset, col, None)
    if hasattr(value, "isoformat"):
        value = value.isoformat()
    return value


async def export_assets(db: AsyncSession, data: ExportAssetsRequest) -> BytesIO:
    query = select(Asset).where(Asset.is_archived == data.is_archived)

    if data.search:
        query = query.where(
            Asset.name.ilike(f"%{data.search}%") | Asset.comment.ilike(f"%{data.search}%")
        )
    if data.organization_id:
        query = query.where(Asset.organization_id == data.organization_id)
    if data.project_id:
        query = query.where(Asset.project_id == data.project_id)
    if data.type_id:
        query = query.where(Asset.type_id == data.type_id)
    if data.currency_id:
        query = query.where(Asset.currency_id == data.currency_id)

    result = await db.execute(query.order_by(Asset.id))
    assets = list(result.scalars().all())

    columns = data.columns or DEFAULT_COLUMNS
    columns = [c for c in columns if c in COLUMN_HEADERS]

    lookups = await _load_lookups(db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Активы"

    # Header row
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=COLUMN_HEADERS.get(col_name, col_name))

    # Data rows
    for row_idx, asset in enumerate(assets, 2):
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=_get_cell_value(asset, col_name, lookups))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
